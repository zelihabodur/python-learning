import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
REPORTS_DIR = os.path.join(PROJECT_DIR, "reports")

excel_path = os.path.join(REPORTS_DIR, "satis_stok_musteri_analiz_raporu.xlsx")

report_files = {
    "KPI Özeti": "yonetici_kpi_ozeti.csv",
    "Şehir Ciro": "sehir_ciro_raporu.csv",
    "Kategori Performans": "kategori_performans_raporu.csv",
    "Müşteri Tipi": "musteri_tipi_raporu.csv",
    "Aylık Ciro": "aylik_ciro_raporu.csv",
    "Kritik Stok": "kritik_stok_raporu.csv",
    "En İyi Ürünler": "en_iyi_urunler_raporu.csv",
    "Ana Satış Verisi": "ana_satis_verisi.csv"
}

with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    for sheet_name, file_name in report_files.items():
        file_path = os.path.join(REPORTS_DIR, file_name)
        df = pd.read_csv(file_path, encoding="utf-8-sig")
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    summary_path = os.path.join(REPORTS_DIR, "yonetici_ozeti.txt")

    with open(summary_path, "r", encoding="utf-8") as file:
        summary_lines = file.read().splitlines()

    summary_df = pd.DataFrame({"Yönetici Özeti": summary_lines})
    summary_df.to_excel(writer, sheet_name="Yönetici Özeti", index=False)


workbook = load_workbook(excel_path)

for worksheet in workbook.worksheets:
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

        adjusted_width = min(max_length + 2, 40)
        worksheet.column_dimensions[column_letter].width = adjusted_width


workbook.save(excel_path)

print("Excel raporu basariyla olusturuldu.")
print(f"- reports/{os.path.basename(excel_path)}")